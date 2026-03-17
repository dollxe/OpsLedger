import datetime
import io
import os
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import (
    TEMPLATE_PATH,
    EXCEL_MAX_SHEETS,
    DONGJU_SHEET, DONGJU_CELL_MAP, DONGJU_BOSUJA_CELL, DONGJU_BOSUJA_NAME,
    SINSEONG_SHEET, SINSEONG_CELL_MAP, SINSEONG_BOSUJA_CELL, SINSEONG_BOSUJA_NAME,
    DATE_CELL,
)
from src.utils import get_weekday_str, get_bosuja, format_sheet_name, is_holiday


def format_b6_date(date: datetime.date) -> str:
    """B6 날짜 셀 문자열 생성. 원본 템플릿 형식: '2024 년  월  일  요일  날씨 :   온도:'"""
    weekday = get_weekday_str(date)
    return (
        f"{date.year} 년\xa0\xa0{date.month}월\xa0\xa0{date.day}일"
        f"\xa0\xa0{weekday}요일\xa0\xa0날씨\xa0:\xa0\xa0\xa0온도:"
    )


def _clear_font_color(cell):
    """셀 폰트 색상을 자동(검정)으로 초기화. 다른 폰트 속성은 유지."""
    f = cell.font
    cell.font = Font(
        name=f.name, size=f.size, bold=f.bold, italic=f.italic,
        underline=f.underline, strike=f.strike,
        # color 생략 → 자동(검정)
    )


def _write_sheet(
    ws,
    date: datetime.date,
    data: dict,
    cell_map: dict,
    bosuja_cell: str,
    facility: str,
):
    """워크시트에 날짜, 데이터, 보수자를 기입."""
    # 날짜
    ws[DATE_CELL] = format_b6_date(date)

    # 전력사용량 / 처리용량 (None이면 빈 값으로 덮어써서 템플릿 기본값 제거)
    for col_key, cells in cell_map.items():
        power_val = data.get(f"{col_key}_power")
        volume_val = data.get(f"{col_key}_volume")
        ws[cells["power"]] = power_val
        ws[cells["volume"]] = volume_val
        # 템플릿의 빨간 폰트 색상 초기화
        _clear_font_color(ws[cells["power"]])
        _clear_font_color(ws[cells["volume"]])

    # 보수자 (월/수/금=이름, 화/목='')
    ws[bosuja_cell] = get_bosuja(date, facility)


def build_workbook(
    template_sheet_name: str,
    cell_map: dict,
    bosuja_cell: str,
    facility: str,
    readings: list[dict],
) -> io.BytesIO:
    """
    날짜별 시트가 담긴 워크북을 BytesIO로 반환.
    템플릿 워크북 내에서 copy_worksheet()로 복사 → 모든 서식 완벽 보존.
    readings: [{'reading_date': 'YYYY-MM-DD', 'dong_bunche_power': ..., ...}]
    """
    if len(readings) > EXCEL_MAX_SHEETS:
        raise ValueError(
            f"시트 수({len(readings)})가 Excel 최대({EXCEL_MAX_SHEETS})를 초과합니다."
        )

    # 템플릿 로드 (매번 새로 로드하여 오염 방지)
    wb = load_workbook(TEMPLATE_PATH)
    template_ws = wb[template_sheet_name]

    # 연도가 여러 해에 걸치는지 확인
    years = set(datetime.date.fromisoformat(r["reading_date"]).year for r in readings)
    include_year = len(years) > 1

    new_sheet_names = []

    for reading in readings:
        date = datetime.date.fromisoformat(reading["reading_date"])
        sheet_name = format_sheet_name(date, include_year=include_year)

        # 중복 시트명 방지
        if sheet_name in wb.sheetnames:
            sheet_name = f"{sheet_name}_{date.strftime('%Y')}"

        # 같은 워크북 내에서 copy_worksheet() → 모든 서식/페이지설정 완벽 복사
        new_ws = wb.copy_worksheet(template_ws)
        new_ws.title = sheet_name
        new_ws.print_area = "B1:AB50"  # 시트명 변경 후 print_area 재설정
        new_ws.print_options.horizontalCentered = True  # 가로 가운데 맞춤
        new_sheet_names.append(sheet_name)

        # 데이터 기입
        _write_sheet(new_ws, date, reading, cell_map, bosuja_cell, facility)

    # 원본 템플릿 시트 및 불필요한 시트 제거 (날짜 시트만 남김)
    for sheet in list(wb.sheetnames):
        if sheet not in new_sheet_names:
            del wb[sheet]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_dongju(readings: list[dict]) -> io.BytesIO:
    return build_workbook(
        template_sheet_name=DONGJU_SHEET,
        cell_map=DONGJU_CELL_MAP,
        bosuja_cell=DONGJU_BOSUJA_CELL,
        facility="dongju",
        readings=readings,
    )


def export_sinseong(readings: list[dict]) -> io.BytesIO:
    return build_workbook(
        template_sheet_name=SINSEONG_SHEET,
        cell_map=SINSEONG_CELL_MAP,
        bosuja_cell=SINSEONG_BOSUJA_CELL,
        facility="sinseong",
        readings=readings,
    )
